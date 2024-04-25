import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl
from email.mime.application import MIMEApplication
from email.encoders import encode_base64
import ollama
import re
import time
import subprocess


def envoyer_email(destinataire, sujet, expediteur, username, mot_de_passe, company):
    # Créer un objet MIMEMultipart pour le message
    prompt_finale = prompt(destinataire, company)
    message = MIMEMultipart()
    message['From'] = expediteur
    message['To'] = destinataire
    text = ""
    text = generation_ollama(prompt_finale)
    subject = re.search(r"Subject:(.*)", text)
    subject = subject.group(1).strip()
    message['Subject'] = subject
    text = re.sub(r"Subject:(.*)", "", text)
    text = re.sub(r"\n\n", "", text,1)


    # Ajouter le contenu du message
    message.attach(MIMEText(text, 'plain'))
    
    message.attach(contenu())

    suspendre_programme()
    
    # Établir une connexion avec le serveur SMTP de Gmail
    serveur_smtp = smtplib.SMTP('smtp.gmail.com', 587)
    serveur_smtp.starttls()

    # Se connecter au compte Gmail
    serveur_smtp.login(username, mot_de_passe)

    # Envoyer le message
    serveur_smtp.send_message(message)

    # Fermer la connexion
    serveur_smtp.quit()
    
def envoyer_email_classeur():
    # Ouvrir le fichier de contact en .xlsx
    classeur = openpyxl.load_workbook('Classeur1.xlsx')
    feuille = classeur.active

    # Récupérer le contenu du message à partir de la feuille de calcul
    adresse_mail = []
    for row in feuille.iter_rows(min_row=4, max_row=5, min_col=6, max_col=6):
        for cell in row:
            adresse_mail.append(cell.value + '\n')
    nom = []
    for row in feuille.iter_rows(min_row=4, max_row=5, min_col=2, max_col=2):
        for cell in row:
            nom.append(cell.value + '\n')
    
    entreprise = []
    for row in feuille.iter_rows(min_row=4, max_row=5, min_col=3, max_col=3):
        for cell in row:
            entreprise.append(cell.value + '\n')
    for i in adresse_mail:
        # Appeler la fonction envoyer_email avec les paramètres appropriés
        destination = 'candidature stage ' + nom[adresse_mail.index(i)]
        company = entreprise[adresse_mail.index(i)]
        envoyer_email(i, destination,"Yourname", "YourEmail", 'yourpassword', company)

def contenu():
    # Créer un objet MIMEMultipart pour le contenu du message
    contenu = MIMEMultipart()

    # Créer les objets MIMEApplication pour les fichiers PDF
    #with open("Yourname.pdf", "rb") as opened:
        #openedfile = opened.read()
    with open("Copie de Yourname.pdf", "rb") as opened2:
        openedfile2 = opened2.read()
    
    #attachedfile = MIMEApplication(openedfile, _subtype="pdf", _encoder=encode_base64)
    attachedfile2 = MIMEApplication(openedfile2, _subtype="pdf", _encoder=encode_base64)
    #attachedfile.add_header('content-disposition', 'attachment', filename="Yourname.pdf")
    attachedfile2.add_header('content-disposition', 'attachment', filename="Yourname EN.pdf")
    #contenu.attach(attachedfile)
    contenu.attach(attachedfile2)
    
    return contenu

def prompt(name, company):
    text = "" # Ajout de la prompt de recherche de stage 
    return text

def extract(output):
    text = ""
    text += output['response']
    return text

def generation_ollama(prompt):
    client = ollama.Client()
    output = client.generate('mistral', prompt)
    return (extract(output))
def suspendre_programme():
    while True:
        try:
            # Vérifier la connexion internet en effectuant un ping vers google.com
            ping_result = subprocess.run(['ping', '-n', '1', 'google.com'], capture_output=True)
            try:
                ping_time = int(ping_result.stdout.decode('utf-8').split('time=')[1].split('ms')[0].strip())
            except UnicodeDecodeError:
                # Handle the UnicodeDecodeError here
                ping_time = 0  # Set a default value or handle the error in a different way
            if ping_time < 200:
                break  # Sortir de la boucle si le ping est inférieur à 200 ms
        except subprocess.CalledProcessError:
            pass  # Ignorer les erreurs de ping

        # Attendre pendant 1 seconde avant de réessayer
        time.sleep(1)
envoyer_email_classeur()
